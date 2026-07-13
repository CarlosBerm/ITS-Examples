"""
get_prompts.py — load the prompt library from the LOCAL eval workbook into a pandas DataFrame.

Local-file twin of GSheets Evaluation/get_prompts.py. Instead of the Google Sheets API + a
service account, it reads the prompt library straight from a local .xlsx workbook with
openpyxl — no Google Cloud, no network, just a file on disk. The table's shape is otherwise
identical, because it IS the same eval database: the 'Prompt Repository' tab's header row is
row 5 and spans columns A–F:

    A: Prompt ID
    B: Core Category (auto)
    C: Subcategory
    D: Context
    E: Prompt Text
    F: Ground Truth / Evaluation Standard

Only rows whose 'Prompt Text' cell is non-empty are kept — that's the prompt that gets sent
to the AI being evaluated.

Configure in Local Evaluation/.env:
    LOCAL_WORKBOOK   path to the local .xlsx eval database
                     (default: '../Evaluation Framework Database Ver0.2.xlsx', the snapshot
                      committed next to the two mode folders)

This module also owns the shared workbook helpers (resolve_workbook_path, open_workbook,
_read_tab_dataframe, _find_column) that get_model_directory.py and populate_model_tab.py
reuse — so the file-opening logic lives in exactly one place, mirroring how the GSheets
scripts share get_sheets_service.
"""

import os
import sys
import warnings

from dotenv import load_dotenv
import openpyxl
import pandas as pd

# Set the current working directory to this file's folder so the .env file and the default
# workbook path resolve the same way no matter where the script is run from.
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# Load configuration (the workbook path) from the .env file.
if not load_dotenv(".env"):
    print("Unable to load .env file.", file=sys.stderr)
    sys.exit(1)

# The eval database lives one level up — the "local snapshot of the eval database" committed
# next to the GSheets/Local mode folders. Override with LOCAL_WORKBOOK in .env.
DEFAULT_WORKBOOK = os.path.join("..", "Evaluation Framework Database Ver0.2.xlsx")
WORKBOOK_PATH = os.environ.get("LOCAL_WORKBOOK", DEFAULT_WORKBOOK)

# The tables' header row is row 5 (same convention as the shared Google Sheet).
HEADER_ROW = 5

# The Prompt Repository table spans columns A–F (6 columns).
PROMPT_LAST_COLUMN = 6

# The column whose emptiness decides whether a row is a real prompt.
PROMPT_TEXT_COLUMN = "Prompt Text"


def resolve_workbook_path(workbook_path=None):
    """Return the workbook path to use (argument, else env LOCAL_WORKBOOK), verifying it exists."""
    path = workbook_path or WORKBOOK_PATH
    if not path:
        print("Error: LOCAL_WORKBOOK is not set in .env.", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(path):
        print(
            f"Error: local workbook '{path}' not found. Point LOCAL_WORKBOOK in .env at your "
            ".xlsx eval database (relative paths are resolved from this folder).",
            file=sys.stderr,
        )
        sys.exit(1)
    return path


def open_workbook(workbook_path=None, data_only=True, read_only=False):
    """Load the local workbook with openpyxl.

    data_only=True returns the values Excel/Sheets last computed for formula cells (what the
    read stages want); data_only=False keeps the formulas themselves intact (what the write
    stage needs so it never clobbers the auto-computed columns). The load-time warning about
    an unsupported 'data validation extension' is suppressed on purpose — one advanced
    cross-sheet dropdown doesn't survive an openpyxl round-trip, but values, formulas, and
    the ordinary dropdowns all do.
    """
    path = resolve_workbook_path(workbook_path)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return openpyxl.load_workbook(path, data_only=data_only, read_only=read_only)
    except Exception as e:
        print(
            f"Error: couldn't open '{path}' as an .xlsx workbook — {type(e).__name__}: {e}",
            file=sys.stderr,
        )
        sys.exit(1)


def _find_column(columns, target):
    """Return the actual column name matching `target`, ignoring case/whitespace."""
    target_norm = target.strip().lower()
    for column in columns:
        if str(column).strip().lower() == target_norm:
            return column
    return None


def _read_tab_dataframe(wb, tab_name, last_col):
    """Read `tab_name` into a DataFrame using sheet row 5 as the header, columns A..last_col.

    openpyxl already returns rectangular rows (empty cells come back as None, padded out to
    last_col), so — unlike the Sheets API reader — there are no ragged rows to normalize.
    Every value is coerced to a string with None -> "", matching what the GSheets reader gets
    back from the API.
    """
    tab = (tab_name or "").strip()
    if tab not in wb.sheetnames:
        available = ", ".join(wb.sheetnames) or "(none)"
        print(
            f"Error: no tab named '{tab_name}' in the workbook. Available tabs: {available}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[tab]
    rows = list(ws.iter_rows(min_row=HEADER_ROW, max_col=last_col, values_only=True))
    if not rows:
        print(f"No data found on tab '{tab}' at or after row {HEADER_ROW}.", file=sys.stderr)
        sys.exit(1)

    def as_text(value):
        return "" if value is None else str(value)

    header = [as_text(v) for v in rows[0]]
    data_rows = [[as_text(v) for v in row] for row in rows[1:]]
    return pd.DataFrame(data_rows, columns=header)


def get_prompts_df(tab_name, workbook_path=None):
    """Read the prompt library from `tab_name` in the local workbook and return it as a
    DataFrame. Sheet row 5 is the header; only rows with a non-empty 'Prompt Text' are kept.
    `workbook_path` defaults to env LOCAL_WORKBOOK.
    """
    wb = open_workbook(workbook_path, data_only=True, read_only=True)
    try:
        df = _read_tab_dataframe(wb, tab_name, PROMPT_LAST_COLUMN)
    finally:
        wb.close()

    prompt_col = _find_column(df.columns, PROMPT_TEXT_COLUMN)
    if prompt_col is None:
        print(
            f"Error: expected a '{PROMPT_TEXT_COLUMN}' column. Found: {list(df.columns)}",
            file=sys.stderr,
        )
        sys.exit(1)

    # Keep only rows whose Prompt Text is non-empty (ignoring surrounding whitespace).
    keep = df[prompt_col].astype(str).str.strip() != ""
    return df[keep].reset_index(drop=True)


def main():
    tab_name = os.environ.get("PROMPT_REPOSITORY_TAB", "Prompt Repository").strip()
    df = get_prompts_df(tab_name)
    print(f"Loaded {len(df)} prompt(s) from '{tab_name}'.\n")
    print(df.to_string(index=False))


if __name__ == "__main__":
    main()
