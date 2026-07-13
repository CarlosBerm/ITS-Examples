"""
populate_model_tab.py — write model responses back into a LOCAL workbook's results tab.

Local-file twin of GSheets Evaluation/populate_model_tab.py. Takes a responses DataFrame (the
get_responses.py format: 'Prompt ID' and 'Model Output Response' columns) and writes those two
columns, row by row, into the matching columns of a results tab in the local .xlsx — then saves
the workbook. Every OTHER column on that tab (Run ID, Date, Model Name, Core/Subcategory, Score,
...) autogenerates or is filled by evaluators, so this script touches ONLY the two columns it's
given and leaves the rest — formulas included — alone.

Behavior notes (matching the GSheets writer):
  * Finds its two target columns by HEADER NAME in row 5, so it writes the right columns even
    if the tab's layout differs from another tab's.
  * Overwrites in place starting at the first data row (row 6, since the header is row 5). It
    does not append below existing data, and it clears any stale rows left over from a longer
    previous run — in those two columns only.
  * Stores every response as literal text: a response that happens to start with '=' (or '+',
    '-', '@') is forced to a string cell so openpyxl never turns it into a live formula. This is
    the local equivalent of the Sheets API's valueInputOption="RAW".
  * Sets the workbook to recalculate on load, so the auto-computed columns and dashboards
    refresh the next time the file is opened in Excel / LibreOffice / Google Sheets.

Reuses get_prompts' resolve_workbook_path / open_workbook so the workbook path (LOCAL_WORKBOOK)
and the load behavior stay in one place. Unlike the read stages it opens the workbook with
data_only=False (to preserve formulas) and writes it back.

Public entry point: populate_model_tab(responses_df, tab_name, workbook_path=None).
"""

import sys

from openpyxl.utils import get_column_letter

from get_prompts import HEADER_ROW, open_workbook, resolve_workbook_path

# The first data row of a results tab (the header is row 5).
FIRST_DATA_ROW = HEADER_ROW + 1

# Column names shared between the responses DataFrame and the results tab's header.
PROMPT_ID_COLUMN = "Prompt ID"
RESPONSE_COLUMN = "Model Output Response"

# Leading characters a spreadsheet would treat as the start of a formula. A value starting
# with any of these is stored as literal text instead of being evaluated.
FORMULA_LEADERS = ("=", "+", "-", "@")


def _set_literal(ws, row, col_idx, text):
    """Write `text` into (row, col_idx) as a literal string — never as a formula.

    openpyxl treats any assigned string starting with '=' as a formula. Forcing the cell's
    data type to 's' (string) keeps a response like '=SUM(...)' stored verbatim.
    """
    cell = ws.cell(row=row, column=col_idx)
    cell.value = text
    if isinstance(text, str) and text.startswith(FORMULA_LEADERS):
        cell.data_type = "s"
    return cell


def _find_header_index(ws, target):
    """Return the 1-based column index of `target` in the tab's header row (row 5),
    ignoring case/whitespace; None if not found."""
    target_norm = target.strip().lower()
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value is not None and str(value).strip().lower() == target_norm:
            return col
    return None


def populate_model_tab(responses_df, tab_name, workbook_path=None):
    """Write the Prompt ID and Model Output Response columns of `responses_df` into `tab_name`
    of the local workbook, overwriting any existing values in those two columns, and save.

    Args:
        responses_df: DataFrame with 'Prompt ID' and 'Model Output Response' columns.
        tab_name: the results tab (worksheet) to write into.
        workbook_path: workbook to write to; defaults to env LOCAL_WORKBOOK.
    """
    # Make sure the DataFrame has what we need before touching the workbook.
    for column in (PROMPT_ID_COLUMN, RESPONSE_COLUMN):
        if column not in responses_df.columns:
            print(
                f"Error: the responses DataFrame is missing a '{column}' column. "
                f"Found: {list(responses_df.columns)}",
                file=sys.stderr,
            )
            sys.exit(1)

    path = resolve_workbook_path(workbook_path)
    # data_only=False so the tab's auto-computed formulas are preserved on save.
    wb = open_workbook(path, data_only=False, read_only=False)

    if tab_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames) or "(none)"
        print(
            f"Error: no tab named '{tab_name}' in '{path}'. Available tabs: {available}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[tab_name]

    # Locate the two target columns by header name, so we write into the right columns and
    # never touch the autogenerate columns.
    prompt_idx = _find_header_index(ws, PROMPT_ID_COLUMN)
    response_idx = _find_header_index(ws, RESPONSE_COLUMN)
    missing = [
        name
        for name, idx in ((PROMPT_ID_COLUMN, prompt_idx), (RESPONSE_COLUMN, response_idx))
        if idx is None
    ]
    if missing:
        header = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
        print(
            f"Error: couldn't find column(s) {missing} in row {HEADER_ROW} of '{tab_name}'. "
            f"Header was: {header}",
            file=sys.stderr,
        )
        sys.exit(1)

    prompt_ids = responses_df[PROMPT_ID_COLUMN].fillna("").astype(str).tolist()
    responses = responses_df[RESPONSE_COLUMN].fillna("").astype(str).tolist()
    n = len(responses_df)

    # Overwrite the two columns in place, storing text literally (no formula evaluation).
    for i in range(n):
        row = FIRST_DATA_ROW + i
        _set_literal(ws, row, prompt_idx, prompt_ids[i])
        _set_literal(ws, row, response_idx, responses[i])

    # Clear any leftover rows below the new data (a shorter run than last time), in just these
    # two columns, so this overwrites rather than leaving stale rows behind. Formulas and the
    # autogenerate columns elsewhere on the row are left untouched.
    for row in range(FIRST_DATA_ROW + n, ws.max_row + 1):
        ws.cell(row=row, column=prompt_idx).value = None
        ws.cell(row=row, column=response_idx).value = None

    # Force Excel / LibreOffice / Google Sheets to recompute the auto columns and dashboards
    # when the file is next opened (openpyxl itself does not evaluate formulas).
    wb.calculation.fullCalcOnLoad = True

    try:
        wb.save(path)
    except PermissionError:
        print(
            f"Error: couldn't save '{path}'. Is it open in Excel? Close the file and re-run.",
            file=sys.stderr,
        )
        sys.exit(1)
    except Exception as e:
        print(f"Error while saving '{path}': {type(e).__name__}: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        wb.close()

    prompt_col = get_column_letter(prompt_idx)
    response_col = get_column_letter(response_idx)
    last_data_row = HEADER_ROW + n
    print(
        f"Wrote {n} row(s) into '{tab_name}' "
        f"(columns {prompt_col}='{PROMPT_ID_COLUMN}', {response_col}='{RESPONSE_COLUMN}', "
        f"rows {FIRST_DATA_ROW}-{last_data_row}) and saved '{path}'."
    )
